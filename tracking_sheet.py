import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from datetime import date

# Class for a new tracking sheet
class Sheet:
    def __init__(self):
        try:
            self.wb = load_workbook(filename="Tracking_Sheet.xlsx")
            self.name = "File 1"
        except:
            workbook = Workbook()
            self.wb = workbook
            sheet = workbook.active
            sheet["A1"] = "Month"
            sheet["B1"] = "Date"
            sheet["C1"] = "Description"
            sheet["D1"] = "Category"
            sheet["E1"] = "Income"
            sheet["F1"] = "Expense"
            sheet["G1"] = "Balance"
            sheet["H1"] = date.today().strftime("%m/%d/%y")
            workbook.save(filename="Tracking_Sheet.xlsx")
            self.name = "File 1"

    def addExpense(self, month, date, description, category, income, expense, amount):
        # Must iterate through each respective column and find the first open slot to put item in *All on the same row
        # Must iterate to find the first open row
        openRow = 0
        rowFound = False
        # Application allows for up to 500 possible values
        for row_cells in self.wb.active.iter_rows(min_row=1, max_row=500, min_col=1, max_col=1):
            for cell in row_cells:
                if cell.value is None:
                    openRow = cell.row
                    print(openRow)
                    rowFound = True
                    break
            if (rowFound):
                break

        self.wb.active.cell(openRow, 1).value = month
        self.wb.active.cell(openRow, 2).value = date
        self.wb.active.cell(openRow, 3).value = description
        self.wb.active.cell(openRow, 4).value = category

        if (income == True):
            self.wb.active.cell(openRow, 5).value = amount
            if (self.wb.active.cell(openRow-1, 5).value == "Income"):
                self.wb.active.cell(openRow, 7).value = amount

            else:
                balance = self.wb.active.cell(openRow-1, 7).value
                self.wb.active.cell(openRow, 7).value = balance + amount


        if (expense == True):
            self.wb.active.cell(openRow, 6).value = amount
            # If this is the first expense / income, make balance equal to this value * -1
            if (openRow == 2):
                self.wb.active.cell(openRow, 7).value = amount * (-1)
            
            elif (openRow > 2):
                balance = self.wb.active.cell(openRow-1, 7).value
                self.wb.active.cell(openRow, 7).value = balance - amount

        self.wb.save(filename="Tracking_Sheet.xlsx")
    
    def getBalance(self):
        # Iterating through the incomes
        balanceFound = False
        cellNum = 1
        balanceCell = 1
        for row_cells in self.wb.active.iter_rows(min_row=1, max_row=500, min_col=7, max_col=7):
            if (balanceFound):
                break
            for cell in row_cells:
                if cell.value == None:
                    balanceCell = cellNum
                    balanceFound = True
                    break
                cellNum += 1
        if (balanceCell != 2):
            return self.wb.active.cell(balanceCell - 1, 7).value
        else:
            return 0

    def clearExpense(self):
        finalRowFound = False
        for i in range(1, 500):
            print(i)
            if (finalRowFound):
                break
            if (self.wb.active.cell(i + 1, 1).value == None):
                finalRowFound = True
                # The last row is at index i, proceed to remove it
                if (i != 1):
                    self.wb.active.delete_rows(i, 1)
                    self.wb.save(filename="Tracking_Sheet.xlsx")
                    print("Deleted")
                    break
                else:
                    print("Cannot Delete Top Row")
                    break
    
    def getCategorySums(self, income):
        catSums = {}
        # Iterate through expense categories and add up their totals
        for i in range (2, 500):
            if (self.wb.active.cell(i, 4).value == None):
                break
            # If the row is not empty
            else:
                # If it is income
                if (income):
                    # If this is an income category
                    if (self.wb.active.cell(i, 5).value != None):
                        # If this is the first element for this category, must add it to the dictionary
                        if (catSums.get(self.wb.active.cell(i, 4).value) == None):
                            catSums[self.wb.active.cell(i, 4).value] = self.wb.active.cell(i, 5).value
                        else:
                            catSums[self.wb.active.cell(i, 4).value] += self.wb.active.cell(i, 5).value
                else:
                    if (self.wb.active.cell(i, 6).value != None):
                        if (catSums.get(self.wb.active.cell(i, 4).value) == None):
                            catSums[self.wb.active.cell(i, 4).value] = (self.wb.active.cell(i, 6).value)
                        else:
                            catSums[self.wb.active.cell(i, 4).value] += self.wb.active.cell(i, 6).value          
        return catSums

    def resetSheet(self):
        self.wb.active.delete_cols(1, 7)
        self.wb.active["A1"] = "Month"
        self.wb.active["B1"] = "Date"
        self.wb.active["C1"] = "Description"
        self.wb.active["D1"] = "Category"
        self.wb.active["E1"] = "Income"
        self.wb.active["F1"] = "Expense"
        self.wb.active["G1"] = "Balance"
        self.wb.active["H1"] = date.today().strftime("%m/%d/%y")
        self.wb.save(filename="Tracking_Sheet.xlsx")

    def getStartDate(self):
        return self.wb.active["H1"].value

    def getMonthlyTotals(self):
        monthTotals = dict()
        for i in range (2, 500):
            if (self.wb.active.cell(i, 1).value == None):
                break
            # Column 5 is income, 6 is expenses
            else:
                # If it is income
                if (self.wb.active.cell(i, 5).value != None):
                    # If the month does not have any values yet
                    if (monthTotals.get(self.wb.active.cell(i, 1).value) == None):
                        monthTotals[self.wb.active.cell(i, 1).value] = [self.wb.active.cell(i, 5).value, 0]
                    else:
                        monthTotals[self.wb.active.cell(i, 1).value][0] += self.wb.active.cell(i, 5).value
                # If it is an expense
                elif(self.wb.active.cell(i, 6).value != None):
                    # If the month has no values yet
                    if (monthTotals.get(self.wb.active.cell(i, 1).value) == None):
                        monthTotals[self.wb.active.cell(i, 1).value] = [0, self.wb.active.cell(i, 6).value]
                    else:
                        monthTotals[self.wb.active.cell(i, 1).value][1] += self.wb.active.cell(i, 6).value
        return monthTotals